#!/usr/bin/env python3

import datetime
import re
import tempfile
from bs4 import BeautifulSoup
import openpyxl
import scrape_common as sc


base_url = 'https://www.ag.ch'
url = f'{base_url}/de/themen_1/coronavirus_2/lagebulletins/lagebulletins_1.jsp'
d = sc.download(url)
soup = BeautifulSoup(d, 'html.parser')

link = soup.find('a', href=re.compile(r'.xlsx$'))
xls_url = link.get('href')
if not xls_url.startswith('http'):
    xls_url = f'{base_url}{xls_url}'
xls_data = sc.download_data(xls_url)

fp = tempfile.NamedTemporaryFile(suffix='.xlsx')
fp.write(xls_data)
book = openpyxl.load_workbook(fp.name)
sheet = book['6. Impfkampagne']

title_row = 3
date_column = 1

first_dose_column = None
second_dose_column = None
total_vaccinations_column = None
doses_delivered_column = None
for column in range(2, sheet.max_column):
    value = sheet.cell(title_row, column).value
    if re.match(r'(Total)\s+Erstimpfungen\s+\(kumuliert\)', value):
        first_dose_column = column
    if re.match(r'(Total)\s+Zweitimpfungen\s+\(kumuliert\)', value):
        second_dose_column = column
    if re.match(r'(Total)\s+Impfungen\s+\(kumuliert\)', value):
        total_vaccinations_column = column
    if re.match(r'(Total)\s+gelieferte\s+Impfdosen\s+\(kumuliert\)', value):
        doses_delivered_column = column

assert first_dose_column
assert second_dose_column
assert total_vaccinations_column
assert doses_delivered_column

for row in range(title_row + 1, sheet.max_row):
    value = sheet.cell(row, first_dose_column).value
    if value is None or value == '':
        continue

    vd = sc.VaccinationData(canton='AG', url=url)

    vd.date = sheet.cell(row, date_column).value.date().isoformat()
    vd.total_vaccinations = int(sheet.cell(row, total_vaccinations_column).value)
    vd.first_doses = int(sheet.cell(row, first_dose_column).value)
    vd.second_doses = int(sheet.cell(row, second_dose_column).value)
    vd.doses_delivered = int(sheet.cell(row, doses_delivered_column).value)

    assert vd
    print(vd)
