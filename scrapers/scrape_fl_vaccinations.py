#!/usr/bin/env python3

import datetime
import tempfile
import openpyxl
import scrape_common as sc


url = 'https://www.llv.li/inhalt/118804/amtsstellen/sonderseite-covid-19'
xls_url = 'https://www.llv.li/files/as/impfungen.xlsx'
xls_data = sc.download_data(xls_url)

fp = tempfile.NamedTemporaryFile(suffix='.xlsx')
fp.write(xls_data)

book = openpyxl.load_workbook(fp.name)
sheet = book['Impfungen']

for i in range(2, sheet.max_column):
    vd = sc.VaccinationData(canton='FL', url=xls_url)
    date = sheet.cell(4, i)
    if date.value is None:
        continue
    vd.date = date.value.date().isoformat()
    vd.doses_delivered = int(sheet.cell(5, i).value)
    vd.total_vaccinations = int(sheet.cell(6, i).value)
    vd.second_doses = int(sheet.cell(7, i).value)
    vd.first_doses = vd.total_vaccinations - vd.second_doses

    assert vd
    print(vd)
