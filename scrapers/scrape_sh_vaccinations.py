#!/usr/bin/env python3

import datetime
import tempfile
import re
import json
import arrow
import openpyxl
from bs4 import BeautifulSoup
import scrape_common as sc


main_url = 'https://coviddashboard.sh.ch/impfungen/'
content = sc.download(main_url)
soup = BeautifulSoup(content, 'html.parser')
url = soup.find('a', href=re.compile(r'.*\.xlsx')).get('href')

xls_data = sc.download_data(url)
fp = tempfile.NamedTemporaryFile(suffix='.xlsx')
fp.write(xls_data)


book = openpyxl.load_workbook(fp.name)
sheet = book['Datensatz_Zeitstrahl']

start_row = 2
date_column = 1
first_dose_column = 12
second_dose_column = 13

for row in range(start_row, sheet.max_row):
    value = sheet.cell(row, first_dose_column).value
    if value is None or value == '':
        continue

    try:
        vd = sc.VaccinationData(canton='SH', url=main_url)

        date = sheet.cell(row, date_column)
        vd.date = date.value.date().isoformat()

        vd.first_doses = int(sheet.cell(row, first_dose_column).value)
        vd.second_doses = int(sheet.cell(row, second_dose_column).value)
        vd.total_vaccinations = vd.first_doses + vd.second_doses

        assert vd
        print(vd)
    except:
        # TODO?
        pass
