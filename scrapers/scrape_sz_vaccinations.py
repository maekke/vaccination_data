#!/usr/bin/env python3

import datetime
import re
import tempfile
import arrow
from bs4 import BeautifulSoup
import openpyxl
import scrape_common as sc


url = 'https://www.sz.ch/behoerden/information-medien/medienmitteilungen/coronavirus.html/72-416-412-1379-6948'
d = sc.download(url)
soup = BeautifulSoup(d, 'html.parser')

link = soup.find('a', text=re.compile(r'Kennzahlen Impfungen'))
xls_url = link.get('href')
xls_data = sc.download_data(xls_url)

fp = tempfile.NamedTemporaryFile(suffix='.xlsx')
fp.write(xls_data)
book = openpyxl.load_workbook(fp.name)
sheet = book['Tabelle1']

title_row = 4
date_column = 1
vaccinations_column = 3

for row in range(title_row + 1, sheet.max_row):
    vd = sc.VaccinationData(canton='SZ', url=url)

    date = sheet.cell(row, date_column).value
    vd.date = date.date().isoformat()

    vd.total_vaccinations = sheet.cell(row, vaccinations_column).value

    assert vd
    print(vd)
